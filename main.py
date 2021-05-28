import math
import pyautogui
import time
import pydirectinput
import random
import cv2 as cv
import os
import openpyxl
from playsound import playsound
import keyboard
import numpy
import sandesh
import os
from slack_sdk import WebClient

def newest(path):
    files = os.listdir(path)
    paths = [os.path.join(path, basename) for basename in files]
    return max(paths, key=os.path.getctime)


def CzyIstnieje(nazwa1):
    global nazwa, wb
    nazwa = nazwa1
    if os.path.exists(nazwa1)==True:
        wb = openpyxl.load_workbook(nazwa1)
    else:
        wb = openpyxl.Workbook()

class Bot():

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.ilosc = 0
        self.wygrane = 0
        self.kart = 0
        self.test = 0
        self.pkt = {}
        self.f = 0
        self.flaga = 0
        self.brylka = 0
        self.wyrzucone = []
        self.brylki = None
        self.Zacznamy()
        self.client = WebClient(token="x")
    
    def Prolog(self):
        # ZMIANA
        time.sleep(0.3)
        if keyboard.is_pressed("End"):
            keyboard.wait("Home")
        cc = time.time()
        self.karty = None
        while self.karty == None:
            self.karty = pyautogui.locateOnScreen("Karty.png", confidence=0.8)
            # ZMIANA
            time.sleep(0.01)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ1")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        pydirectinput.click(self.karty[0] + int(self.karty[2]/2), self.karty[1] + int(self.karty[3]/2))
        self.start = None
        while self.start == None:
            self.start = pyautogui.locateOnScreen("Start.png", confidence=0.8)
            # ZMIANA
            time.sleep(0.01)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ2")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        pydirectinput.click(self.start[0] + int(self.start[2] / 2), self.start[1] + int(self.start[3] / 2))
        self.tak1 = None
        while self.tak1 == None:
            self.tak1 = pyautogui.locateOnScreen("Tak.png", confidence=0.8)
            # ZMIANA
            time.sleep(0.01)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ3")
                mp3Fsiile = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        pydirectinput.click(self.tak1[0] + int(self.tak1[2] / 2), self.tak1[1] + int(self.tak1[3] / 2))
        self.okno = None
        while self.okno == None:
            self.okno = pyautogui.locateOnScreen("Okno.png", confidence=0.8)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ4")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        self.dodawacz = None
        while self.dodawacz  == None:
            self.dodawacz  = pyautogui.locateOnScreen("Dodawacz.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ5")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        self.komorki = None
        while self.komorki  == None:
            self.komorki  = pyautogui.locateOnScreen("Komorki.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        #print(time.time() - czas2)
        self.zakoncz= None
        while self.zakoncz  == None:
            self.zakoncz  = pyautogui.locateOnScreen("Zakoncz.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            # ZMAIANA
            time.sleep(0.01)
            if time.time() - cc>= 20:
                print("PROLOGJEJOJ")
                mp3File = "JD.mp3"
                playsound(mp3File)
                keyboard.wait("HOME")
                print("KONEIC")
        if keyboard.is_pressed("End"):
            keyboard.wait("Home")

    def Prolog2(self):
        time.sleep(0.3)
        sprawdzam = time.time()
        if keyboard.is_pressed("End"):
            keyboard.wait("Home")
        cc = time.time()
        self.start = None
        while self.start == None:
            pydirectinput.click(self.karty[0] + int(self.karty[2] / 2), self.karty[1] + int(self.karty[3] / 2))
            self.start = pyautogui.locateOnScreen("Start.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            if time.time() - cc >= 7:
                pydirectinput.keyDown('printscreen')
                time.sleep(0.2)
                pydirectinput.keyUp('printscreen')
                try:
                    result = self.client.files_upload(
                        channels=channel_id,
                        initial_comment= str("PROLOG " + "ILOSC GIER:" + str(self.ilosc) + " WYGRANE: " + str(self.wygrane) + " MINUT: " + str((time.time() - self.czaaas) / 60)),
                        file=newest("E),
                    )
                except Exception as e:
                    print(e)
                print("PROLOGJEJOJ134")
                mp3File = "JD.mp3"
                playsound(mp3File)
                #keyboard.wait("HOME")
                #print("KONEIC")
        pydirectinput.click(self.start[0] + int(self.start[2] / 2), self.start[1] + int(self.start[3] / 2))
        self.tak1 = None
        while self.tak1 == None:
            self.tak1 = pyautogui.locateOnScreen("Tak.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            if time.time() - cc >= 12:

                pydirectinput.keyDown('printscreen')
                time.sleep(0.2)
                pydirectinput.keyUp('printscreen')
                try:
                    result = self.client.files_upload(
                        channels=channel_id,
                        initial_comment= str("PROLOG " + "ILOSC GIER:" + str(self.ilosc) + " WYGRANE: " + str(self.wygrane) + " MINUT: " + str((time.time() - self.czaaas) / 60)),
                        file=newest(),
                    )
                except Exception as e:
                    print(e)
                print("PROLOGJEJOJ123")
                mp3File = "JD.mp3"
                playsound(mp3File)
                #keyboard.wait("HOME")
        pydirectinput.click(self.tak1[0] + int(self.tak1[2] / 2), self.tak1[1] + int(self.tak1[3] / 2))

        self.komorki = None
        while self.komorki  == None:
            self.komorki  = pyautogui.locateOnScreen("Komorki.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]), confidence=0.8)
            if time.time() - cc>= 15:
                if self.brylka == 0:
                    while self.brylki == None:
                        self.brylki = pyautogui.locateOnScreen("Brylki.png", confidence=0.9)
                    keyboard.press("SHIFT")
                    keyboard.press("ALT")
                    for i in range(25):
                        pydirectinput.click(self.brylki[0] + int(self.brylki[2] / 2), self.brylki[1] + int(self.brylki[3] / 2))
                    keyboard.release("SHIFT")
                    keyboard.release("ALT")
                    self.brylka = 1
                    self.Zacznamy2()
                    return

                pydirectinput.keyDown('printscreen')
                time.sleep(0.2)
                pydirectinput.keyUp('printscreen')
                try:
                    result = self.client.files_upload(
                        channels=channel_id,
                        initial_comment= str("PROLOG " + "ILOSC GIER:" + str(self.ilosc) + " WYGRANE: " + str(self.wygrane) + " MINUT: " + str((time.time() - self.czaaas) / 60)),
                        file=newest("E),
                    )
                except Exception as e:
                    print(e)
                print("PROLOGJEJOJ")
                mp3File = "JD.mp3"
                playsound(mp3File)
                #keyboard.wait("HOME")
                print("KONEIC")
        if self.brylka == 1:
            self.brylka = 0
        if keyboard.is_pressed("End"):
            keyboard.wait("Home")
        print(time.time() - sprawdzam)
        if keyboard.is_pressed("End"):
            keyboard.wait("Home")

    def algorytm(self):
        try:
            response = self.client.chat_postMessage(channel=channel_id, text="Zaczynamy, LECYMY DuRRrrR")
        except Exception as e:
            print(e)
        self.koniec = 2002
        self.skip = 0
        self.czaaas = time.time()
        CzyIstnieje("")
        self.sheet = wb.active
        self.n = 2
        while self.sheet.cell(row=self.n, column=1).value != None:
            self.n += 1
        self.zapisywacz = 0
        while self.ilosc != self.koniec:
            #self.pruba4Z() vs
            self.pruba4ALGTEST()
        print(time.time() - self.czaaas)
        sandesh.send("KONIEC",
                     webhook=)
        print("PROLOGJEJOJ")
        mp3File = "JD.mp3"
        playsound(mp3File)
        keyboard.wait("HOME")
        print("KONEIC")


    def Aktualne(self):
        #lista = []
        im = pyautogui.screenshot(region=(self.komorki[0], self.komorki[1],self.komorki[2],self.komorki[3]))
        #im.save("SS.png")
        #img = cv.imread(f"SS.png", cv.IMREAD_UNCHANGED)
        pil_image = im.convert('RGB')
        open_cv_image = numpy.array(pil_image)
        # Convert RGB to BGR
        img = open_cv_image[:, :, ::-1].copy()

        for i in list(self.pomoc):
            needle_img = cv.imread(f"{i}.png", cv.IMREAD_UNCHANGED)
            result = cv.matchTemplate(img, needle_img, cv.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv.minMaxLoc(result)

            if float(max_val) >= 0.88 and i not in list(self.aktualne):
                self.aktualne[i] = max_loc
                if i not in self.zapis:
                    self.zapis.append(i)
                else:
                    print(i, len(self.zapis),"JUZ jest??", self.zapis)
            #if float(max_val) >= 0.88:
            #    lista.append(i)


    def Zacznamy(self):
        self.czas = time.time()
        self.wyrzucone = []
        self.Prolog()
        self.punkty = 0
        self.aktualne = {}
        self.zapis = []
        self.Pomoc()
        self.flagusia = 0
        for i in range(5):
            self.Dodaj()
            #ZMIANA
            time.sleep(0.01)
        while len(self.aktualne) != 5:
            self.Aktualne()
            self.Dodaj()
        self.Aktualne()

    def Zacznamy2(self):
        self.czas = time.time()
        self.wyrzucone = []
        self.Prolog2()
        self.punkty = 0
        self.aktualne = {}
        self.zapis = []
        self.Pomoc()
        self.flagusia = 0
        for i in range(5):
            self.Dodaj()
            #ZMIANA
            time.sleep(0.01)
        while len(self.aktualne) != 5:
            self.Aktualne()
            self.Dodaj()
        self.Aktualne()

    def pruba4ALGTEST(self):


    def Pomoc(self):
        self.pomoc = []
        for i in range(1,9):
            self.pomoc.append(f"{i}C")
        for i in range(1,9):
            self.pomoc.append(f"{i}N")
        for i in range(1,9):
            self.pomoc.append(f"{i}Z")

    def Usun(self, el):

        self.czasik = time.time()
        a = self.aktualne[el]
        #a = pyautogui.locateOnScreen("6N.png", confidence = 0.88)
        tak = None
        licznik = 0
        while tak == None:
            pydirectinput.rightClick(self.komorki[0] + a[0], self.komorki[1] + a[1])
            #pydirectinput.rightClick(a[0] + int(a[2]/2), a[1] + int(a[3]/2))
            tak = pyautogui.locateOnScreen("Tak.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]),confidence=0.9)
            #tak = pyautogui.locateOnScreen("Tak.png", confidence = 0.8)
            # ZMIANA
            time.sleep(0.01)
            """
            licznik+=1
            if licznik == 10:
                break
            """
            if time.time() - self.czasik >= 15:

                pydirectinput.keyDown('printscreen')
                time.sleep(0.2)
                pydirectinput.keyUp('printscreen')
                try:
                    result = self.client.files_upload(
                        channels=channel_id,
                        initial_comment=str(
                            "USUN " + "ILOSC GIER:" + str(self.ilosc) + " WYGRANE: " + str(
                                self.wygrane) + " MINUT: " + str((time.time() - self.czaaas) / 60)),
                        file=newest(),
                    )
                except Exception as e:
                    print(e)
                print("OJEJ15")
                mp3File = "JD.mp3"
                playsound(mp3File)
                #keyboard.wait("HOME")
                #print("KONEIC")
        pydirectinput.click(tak[0] + int(tak[2]/2), tak[1] + int(tak[3]/2))
        #ZMIANA
        time.sleep(0.03)
        self.Dodaj()

    def Dodaj(self):
        #self.dodawacz = pyautogui.locateOnScreen("Dodawacz.png", confidence = 0.8)
        pydirectinput.click(self.dodawacz[0] + int(self.dodawacz[2]/2), self.dodawacz[1] + int(self.dodawacz[3]/2))

    def DodajDoBinga(self, el):

        self.czasik = time.time()
        a = self.aktualne[el]
        flaga = 0
        licznik = 0
        while flaga == 0:
            pydirectinput.click(self.komorki[0] + a[0], self.komorki[1] + a[1])
            b = pyautogui.locateOnScreen(f"{el}.png",region = (self.okno[0], self.okno[1], self.okno[0] + self.okno[2], self.okno[1] + self.okno[3]),confidence=0.88)
            #ZMIANA
            time.sleep(0.01)
            try:
                if math.fabs(self.komorki[1] + a[1]- b[1]) >= 50:
                    flaga = 1
            except:
                licznik+=1
                if licznik == 2:
                    #print("Nie udana flaga w dodaj bingo")
                    flaga = 1
            if time.time() - self.czasik >= 15:

                pydirectinput.keyDown('printscreen')
                time.sleep(0.2)
                pydirectinput.keyUp('printscreen')
                try:
                    result = self.client.files_upload(
                        channels=channel_id,
                        initial_comment=str(
                            "BINGO " + "ILOSC GIER:" + str(self.ilosc) + " WYGRANE: " + str(
                                self.wygrane) + " MINUT: " + str((time.time() - self.czaaas) / 60)),
                        file=newest(),
                    )
                except Exception as e:
                    print(e)
                print("OJE16J")
                mp3File = "JD.mp3"
                playsound(mp3File)
                #keyboard.wait("HOME")
                #print("KONEIC")

    def algorytm2R(self): #glebokosc 2 ~43%
        punkty = {}
        for i in list(self.aktualne):
            punkty[str(i)] = 0


        for i in self.aktualne:
            ######PRZYGOTOWANIE
            lista1 = list(self.aktualne)
            lista1.remove(i)

            lista2 = list(self.pomoc)
            lista2.remove(i)
            #############

            for j in lista2:
                lista1.append(j)
                if self.bingo(False, lista1) != 0:
                    punkty[i]+= 100

                for q in lista1:

                    lista3 = list(lista1)
                    lista3.remove(q)

                    lista4 = list(lista2)
                    lista4.remove(q)

                    for r in list(lista4):
                        lista3.append(r)
                        if self.bingo(False, lista3) != 0:
                            punkty[i] += 100

                        lista3.remove(r)



                lista1.remove(j)

        min = punkty[list(punkty)[0]]
        zap = []
        for i in list(punkty):
            if punkty[i] == min:
                zap.append(i)
            if punkty[i] > min:
                min = punkty[i]
                zap = []
                zap.append(i)

        if len(zap)!= 1:
            for i in zap:
                if self.rozwbingo(i) != 1:
                    #print(zap, punkty, "1")
                    return i, sorted(punkty, key=punkty.get,reverse=True)
        else:
            #print(zap, punkty, "3")
            return zap[0], sorted(punkty, key=punkty.get,reverse=True)

        #print(zap, punkty, "2")
        cz,n,z = self.prawiebingo()
        for i in zap:
            if i in cz or i in n or i in z:
                zap.remove(i)
        return random.choice(zap), sorted(punkty, key=punkty.get,reverse=True)
        #print(zap, punkty)
        #return sorted(punkty, key=punkty.get,reverse=True)


    def IloscPsucia(self, lista, el):
        lis = list(lista)
        literkilka = []
        for i in lista:
            if i[1] == el[1]:
                literkilka.append(i)
        wynik = []
        licznik = 0
        for i in literkilka:
            if i == el:
                wynik.append(licznik)
                licznik = 0
            else:
                licznik+=1
        wynik.append(licznik)
        return min(wynik)

    def rozwbingo(self, el): #Sprawdzamy czy dany element psuje bingo
        lista = list(self.pomoc)
        dl = len(self.mozliwebinga(lista, False))
        try:
            lista.remove(str(el))
        except:
            print("hij")
            return
        if len(self.mozliwebinga(lista, False)) < dl:
            return 1 #Rozwala bingo
        else:
            return 0 #Nie rozwala binga

    def mozliwebinga(self, lista, odw):
        c = []
        n = []
        z = []
        binga = []
        for i in lista:
            if i[1] == "C":
                c.append(i)
            elif i[1] == "N":
                n.append(i)
            elif i[1] == "Z":
                z.append(i)

        licz = 0
        a = sorted(c, reverse= odw)
        tym = []
        flaga = 0
        for i in range(0, len(a)-1):
            if flaga == 1:
                flaga = 0
                continue
            if math.fabs(int(a[i+1][0]) - int(a[i][0])) == 1:
                licz+=1
                if a[i] not in tym:
                    tym.append(a[i])

                if licz == 2:
                    tym.append(a[i+1])
                    binga.append(tym)
                    tym = []
                    licz = 0
                    flaga = 1
            else:
                licz = 0
                tym = []

        licz = 0
        a = sorted(n, reverse= odw)
        tym = []
        flaga = 0
        for i in range(0, len(a)-1):
            if flaga == 1:
                flaga = 0
                continue
            if math.fabs(int(a[i+1][0]) - int(a[i][0])) == 1:
                licz+=1
                if a[i] not in tym:
                    tym.append(a[i])

                if licz == 2:
                    tym.append(a[i+1])
                    binga.append(tym)
                    tym = []
                    licz = 0
                    flaga = 1
            else:
                licz = 0

        licz = 0
        a = sorted(z, reverse= odw)
        tym = []
        flaga = 0
        for i in range(0, len(a)-1):
            if flaga == 1:
                flaga = 0
                continue
            if math.fabs(int(a[i+1][0]) - int(a[i][0])) == 1:
                licz+=1
                if a[i] not in tym:
                    tym.append(a[i])

                if licz == 2:
                    tym.append(a[i+1])
                    binga.append(tym)
                    tym = []
                    licz = 0
                    flaga = 1
            else:
                licz = 0

        return binga

    def prawiebingo(self):
        c = []
        n = []
        z = []
        gitc = []
        gitn = []
        gitz = []
        for i in list(self.aktualne):
            if i[1] == "C":
                c.append(int(i[0]))
            elif i[1] == "N":
                n.append(int(i[0]))
            elif i[1] == "Z":
                z.append(int(i[0]))
        if len(c) >= 2:
            a = sorted(c)
            for i in range(1, len(a)):
                if a[i] - a[i-1] == 1 or a[i] - a[i-1] == 2:
                    if str(a[i]) + "C" not in gitc:
                        gitc.append(str(a[i]) + "C")
                    if str(a[i-1]) + "C" not in gitc:
                        gitc.append(str(a[i-1]) + "C")
        if len(n) >= 2:
            a = sorted(n)
            for i in range(1, len(a)):
                if a[i] - a[i-1] == 1 or a[i] - a[i-1] == 2:
                    if str(a[i]) + "N" not in gitn:
                        gitn.append(str(a[i]) + "N")
                    if str(a[i-1]) + "N" not in gitn:
                        gitn.append(str(a[i-1]) + "N")
        if len(z) >= 2:
            a = sorted(z)
            for i in range(1, len(a)):
                if a[i] - a[i-1] == 1 or a[i] - a[i-1] == 2:
                    if str(a[i]) + "Z" not in gitz:
                        gitz.append(str(a[i]) + "Z")
                    if str(a[i - 1]) + "Z" not in gitz:
                        gitz.append(str(a[i - 1]) + "Z")
        return gitc, gitn, gitz

    def bingo(self, odw, lista):
        c = []
        n = []
        z = []
        bingo = []
        for i in lista:
            if i[1] == "C":
                c.append(int(i[0]))
            elif i[1] == "N":
                n.append(int(i[0]))
            elif i[1] == "Z":
                z.append(int(i[0]))
        if len(c) >= 3:
            licz = 0
            a = sorted(c, reverse= odw)
            for i in range(1, len(a)):
                if math.fabs(a[i] - a[i - 1]) == 1:
                    licz += 1

                    if a[i] not in bingo:
                        bingo.append(a[i])
                    if a[i - 1] not in bingo:
                        bingo.append(a[i - 1])

                    if licz == 2:
                        return [str(s) + "C" for s in bingo]
                else:
                    licz = 0
                    bingo = []

        elif len(n) >= 3:
            licz = 0
            a = sorted(n, reverse= odw)
            for i in range(1,len(a)):
                if math.fabs(a[i] - a[i - 1]) == 1:
                    licz+= 1

                    if a[i] not in bingo:
                        bingo.append(a[i])
                    if a[i-1] not in bingo:
                        bingo.append(a[i-1])

                    if licz == 2:
                        return [str(s) + "N" for s in bingo]
                else:
                    licz = 0
                    bingo = []
        elif len(z) >= 3:
            licz = 0
            a = sorted(z, reverse= odw)
            for i in range(1, len(a)):
                if math.fabs(a[i] - a[i - 1]) == 1:
                    licz += 1

                    if a[i] not in bingo:
                        bingo.append(a[i])
                    if a[i - 1] not in bingo:
                        bingo.append(a[i - 1])

                    if licz == 2:
                        return [str(s) + "Z" for s in bingo]
                else:
                    licz = 0
                    bingo = []
        return 0

    def BingoPsuje(self, odw): #Czy bingo psuuje z pkt
        lista = []
        for i in list(self.pomoc):
            lista.append(i)
        d1 = len(self.mozliwebinga(lista, False))
        a = self.bingo(odw, list(self.aktualne))
        if a!= 0:
            for i in a:
                lista.remove(i)
        d2 = len(self.mozliwebinga(lista, False))
        if math.fabs(d1 - d2) >= 2 and self.punkty + d2*100 + 100 < 400:
            return 1
        else:
            return 0

    def BingoPsuje2(self, odw): #Czy bingo psuuje bez pkt
        lista = []
        for i in list(self.pomoc):
            lista.append(i)
        d1 = len(self.mozliwebinga(lista, False))

        a = self.bingo(odw, list(self.aktualne))
        if a!= 0:
            for i in a:
                lista.remove(i)
        d2 = len(self.mozliwebinga(lista, False))
        if math.fabs(d1 - d2) >= 2:
            return 1
        else:
            return 0


keyboard.wait("Home")
ss = Bot()
ss.algorytm()
